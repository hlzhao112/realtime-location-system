"""Excel 模板读写与映射表解析。"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .normalize import cell_str, pretty_epc

POS_HEADERS = ["生产线", "生产线ID", "工序", "工序ID", "区域", "区域ID", "蓝牙标签编号"]
NO_HEADERS = ["No.", "EPC"]
TZ_HEADERS = ["台座名称", "编号", "台座安装LYG"]

POS_ALIASES = {
    "生产线": {"生产线", "产线", "产线名称", "line", "line_name"},
    "生产线ID": {"生产线id", "生产线ID", "产线id", "产线ID", "产线编码", "line_id", "lineid"},
    "工序": {"工序", "工序名称", "procedure", "proc"},
    "工序ID": {"工序id", "工序ID", "工序编码", "procedure_id", "proccode"},
    "区域": {"区域", "区域名称", "area", "area_name"},
    "区域ID": {"区域id", "区域ID", "areaid", "area_id", "区域编码"},
    "蓝牙标签编号": {"蓝牙标签编号", "标签编号", "编号", "no.", "no", "tag"},
}
NO_ALIASES = {
    "No.": {"no.", "no", "编号", "标签编号", "蓝牙标签编号"},
    "EPC": {"epc", "epc实际值", "标签实际值"},
}
TZ_ALIASES = {
    "台座名称": {"台座名称", "名称", "台车名称"},
    "编号": {"编号", "台车id", "台车ID", "tz"},
    "台座安装LYG": {"台座安装lyg", "台座安装LYG", "读卡器", "读卡器id", "读卡器ID", "设备id", "设备ID", "reader"},
}

HEADER_FILL = PatternFill("solid", fgColor="E8F6F0")
HINT_FILL = PatternFill("solid", fgColor="F4F4F4")


def _norm_header(v: str) -> str:
    return cell_str(v).lower().replace(" ", "")


def _map_headers(row: list[str], aliases: dict[str, set[str]], required: list[str]) -> dict[str, int]:
    found: dict[str, int] = {}
    used: set[int] = set()
    for i, raw in enumerate(row):
        key = _norm_header(raw)
        if not key:
            continue
        for canon, als in aliases.items():
            als_n = {_norm_header(x) for x in als}
            if key in als_n and canon not in found and i not in used:
                found[canon] = i
                used.add(i)
                break
    # 位置表现场文件第 5、6 列都叫「区域」：第 6 列当区域ID
    if "区域ID" not in found and "区域" in found:
        area_i = found["区域"]
        for i, raw in enumerate(row):
            if i != area_i and _norm_header(raw) in {"区域", "区域id", "区域ID"}:
                found["区域ID"] = i
                break
    missing = [h for h in required if h not in found]
    if missing:
        raise ValueError(f"缺列 / 表头不符：缺少 { '、'.join(missing) }")
    return found


def _iter_rows(ws: Worksheet) -> list[list[Any]]:
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows


def read_first_sheet(data: bytes) -> list[list[Any]]:
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = _iter_rows(ws)
    wb.close()
    if not rows:
        raise ValueError("工作表为空")
    return rows


def parse_position_file(data: bytes) -> tuple[list[dict], list[dict]]:
    """返回 (有效行, 错误)。"""
    rows = read_first_sheet(data)
    headers = [cell_str(x) for x in rows[0]]
    errors: list[dict] = []
    try:
        idx = _map_headers(headers, POS_ALIASES, POS_HEADERS)
    except ValueError as e:
        errors.append({"level": "err", "title": "位置-编号映射表表头", "detail": str(e), "file": "位置表", "row": 1, "column": "表头"})
        return [], errors
    out = []
    for n, row in enumerate(rows[1:], start=2):
        if not any(cell_str(c) for c in row):
            continue
        item = {k: cell_str(row[idx[k]] if idx[k] < len(row) else "") for k in POS_HEADERS}
        empty = [k for k in POS_HEADERS if not item[k]]
        if empty:
            errors.append(
                {
                    "level": "err",
                    "title": "必填单元格为空",
                    "detail": f"位置表 第 {n} 行 · {' / '.join(empty)} 为空",
                    "file": "位置表",
                    "row": n,
                    "column": empty[0],
                }
            )
            continue
        item["_row"] = n
        out.append(item)
    return out, errors


def parse_no_file(data: bytes) -> tuple[dict[str, list[str]], list[dict]]:
    rows = read_first_sheet(data)
    headers = [cell_str(x) for x in rows[0]]
    errors: list[dict] = []
    try:
        idx = _map_headers(headers, NO_ALIASES, NO_HEADERS)
    except ValueError as e:
        errors.append({"level": "err", "title": "编号-映射表表头", "detail": str(e), "file": "编号表", "row": 1, "column": "表头"})
        return {}, errors
    nos: dict[str, list[str]] = {}
    last = ""
    seen_epc: dict[str, str] = {}
    for n, row in enumerate(rows[1:], start=2):
        no = cell_str(row[idx["No."]] if idx["No."] < len(row) else "")
        epc = cell_str(row[idx["EPC"]] if idx["EPC"] < len(row) else "")
        if no:
            last = no
        if not epc:
            if no or any(cell_str(c) for c in row):
                errors.append(
                    {
                        "level": "err",
                        "title": "必填单元格为空",
                        "detail": f"编号表 第 {n} 行 · EPC 为空",
                        "file": "编号表",
                        "row": n,
                        "column": "EPC",
                    }
                )
            continue
        if not last:
            errors.append(
                {
                    "level": "err",
                    "title": "标签编号未定义",
                    "detail": f"编号表 第 {n} 行 · No. 为空且无法延续上一编号",
                    "file": "编号表",
                    "row": n,
                    "column": "No.",
                }
            )
            continue
        from .normalize import norm_epc

        key = norm_epc(epc)
        if key in seen_epc and seen_epc[key] != last:
            errors.append(
                {
                    "level": "err",
                    "title": "EPC 重复",
                    "detail": f"编号表 第 {n} 行 · EPC 与 {seen_epc[key]} 重复",
                    "file": "编号表",
                    "row": n,
                    "column": "EPC",
                }
            )
            continue
        seen_epc[key] = last
        nos.setdefault(last, []).append(pretty_epc(epc) or epc)
    return nos, errors


def parse_trolley_file(data: bytes) -> tuple[list[dict], list[dict]]:
    rows = read_first_sheet(data)
    headers = [cell_str(x) for x in rows[0]]
    errors: list[dict] = []
    try:
        idx = _map_headers(headers, TZ_ALIASES, TZ_HEADERS)
    except ValueError as e:
        errors.append({"level": "err", "title": "台座编号映射表表头", "detail": str(e), "file": "台座表", "row": 1, "column": "表头"})
        return [], errors
    out = []
    seen_tz: dict[str, int] = {}
    seen_rd: dict[str, int] = {}
    for n, row in enumerate(rows[1:], start=2):
        if not any(cell_str(c) for c in row):
            continue
        name = cell_str(row[idx["台座名称"]] if idx["台座名称"] < len(row) else "")
        tz = cell_str(row[idx["编号"]] if idx["编号"] < len(row) else "")
        reader = cell_str(row[idx["台座安装LYG"]] if idx["台座安装LYG"] < len(row) else "")
        if not name or not tz or not reader:
            miss = [k for k, v in (("台座名称", name), ("编号", tz), ("台座安装LYG", reader)) if not v]
            errors.append(
                {
                    "level": "err",
                    "title": "必填单元格为空",
                    "detail": f"第 {n} 行 · {' / '.join(miss)} 为空",
                    "file": "台座表",
                    "row": n,
                    "column": miss[0],
                }
            )
            continue
        if tz in seen_tz:
            errors.append(
                {
                    "level": "err",
                    "title": "台座编号重复",
                    "detail": f"第 {n} 行 · 编号 {tz} 与第 {seen_tz[tz]} 行重复",
                    "file": "台座表",
                    "row": n,
                    "column": "编号",
                }
            )
            continue
        if reader in seen_rd:
            errors.append(
                {
                    "level": "err",
                    "title": "读卡器 ID 重复",
                    "detail": f"第 {n} 行 · 读卡器 {reader} 与第 {seen_rd[reader]} 行重复",
                    "file": "台座表",
                    "row": n,
                    "column": "台座安装LYG",
                }
            )
            continue
        seen_tz[tz] = n
        seen_rd[reader] = n
        out.append({"tz": tz, "name": name, "reader": reader, "_row": n})
    return out, errors


def build_lines_from_pos(pos_rows: list[dict], no_map: dict[str, list[str]]) -> tuple[list[dict], dict[str, list[str]], list[str]]:
    """按位置表出现顺序拼产线树。返回 lines, used_nos, undefined_nos."""
    lines: list[dict] = []
    line_ix: dict[str, dict] = {}
    used: dict[str, list[str]] = {}
    undefined: list[str] = []

    for row in pos_rows:
        lid, lname = row["生产线ID"], row["生产线"]
        if lid not in line_ix:
            node = {"id": lid, "name": lname, "procs": []}
            line_ix[lid] = node
            lines.append(node)
        line = line_ix[lid]
        proc_ix = {p["code"]: p for p in line["procs"]}
        if row["工序ID"] not in proc_ix:
            proc = {"code": row["工序ID"], "name": row["工序"], "order": len(line["procs"]) + 1, "areas": []}
            line["procs"].append(proc)
            proc_ix[row["工序ID"]] = proc
        proc = proc_ix[row["工序ID"]]
        area_ix = {a["id"]: a for a in proc["areas"]}
        if row["区域ID"] not in area_ix:
            area = {"id": row["区域ID"], "name": row["区域"], "nos": [], "epcs": []}
            proc["areas"].append(area)
            area_ix[row["区域ID"]] = area
        area = area_ix[row["区域ID"]]
        no = row["蓝牙标签编号"]
        if no not in area["nos"]:
            area["nos"].append(no)
        if no not in no_map:
            if no not in undefined:
                undefined.append(no)
        else:
            used[no] = no_map[no]
            if not area["epcs"]:
                area["epcs"] = no_map[no][:]
    return lines, used, undefined


def _style_header(ws: Worksheet, headers: list[str], widths: list[int]):
    ws.append(headers)
    for i, w in enumerate(widths, start=1):
        cell = ws.cell(1, i)
        cell.font = Font(bold=True, color="1A1A1A")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = w


def _hint_sheet(wb: Workbook, title: str, lines: list[str]):
    ws = wb.create_sheet(title)
    ws.column_dimensions["A"].width = 88
    ws.append(["填写说明"])
    ws["A1"].font = Font(bold=True)
    for line in lines:
        ws.append([line])
        ws.cell(ws.max_row, 1).fill = HINT_FILL


def build_template(kind: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    if kind == "pos":
        ws.title = "位置-编号映射表"
        _style_header(ws, POS_HEADERS, [14, 14, 16, 12, 20, 12, 16])
        ws.append(["2#产线", "ZCX_570", "混凝土浇筑", "GX_0015", "混凝土浇筑2#区", "QY_19", "B1"])
        ws.append(["2#产线", "ZCX_570", "混凝土浇筑", "GX_0015", "混凝土浇筑2#区", "QY_19", "B2"])
        ws.append(["2#产线", "ZCX_570", "混凝土浇筑", "GX_0015", "混凝土浇筑2#区", "QY_19", "B3"])
        _hint_sheet(
            wb,
            "填写说明",
            [
                "一行 = 一个「区域 + 一个蓝牙标签编号」。同一区域通常连续 3 行（3 个编号）。",
                "工序顺序 = 本文件中该产线工序的出现顺序，导入后仍可在页面拖拽调整。",
                "必须与「编号-映射表」一起导入；系统按标签编号拼接 EPC。",
            ],
        )
    elif kind == "no":
        ws.title = "编号-映射表"
        _style_header(ws, NO_HEADERS, [12, 44])
        ws.append(["B1", "E2 80 68 94 00 00 40 35 80 04 31 0A"])
        ws.append(["", "E2 80 68 94 00 00 50 35 80 04 35 0A"])
        ws.append(["", "E2 80 68 94 00 00 50 35 80 04 2D 0A"])
        _hint_sheet(
            wb,
            "填写说明",
            [
                "同一编号只在首行填写 No.，下方行编号列留空表示延续上一编号。",
                "一个编号可对应多条 EPC；读到任一 EPC 即可定位到该编号所属区域。",
                "位置表引用但本表未定义的编号会阻止导入。",
            ],
        )
    else:
        ws.title = "台座编号映射"
        _style_header(ws, TZ_HEADERS, [18, 14, 16])
        ws.append(["制梁台座#1", "TZ_4710", "LYG1"])
        _hint_sheet(
            wb,
            "填写说明",
            [
                "一行 = 一台台车（台座）与其读卡器的绑定。",
                "编号、读卡器 ID 在本项目内必须唯一。空行自动跳过。",
            ],
        )
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_position(lines: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "当前数据"
    _style_header(ws, POS_HEADERS, [14, 14, 16, 12, 20, 12, 16])
    for line in lines:
        for proc in line.get("procs") or []:
            for area in proc.get("areas") or []:
                nos = area.get("nos") or []
                if not nos:
                    ws.append([line["name"], line["id"], proc["name"], proc["code"], area["name"], area["id"], ""])
                for no in nos:
                    ws.append([line["name"], line["id"], proc["name"], proc["code"], area["name"], area["id"], no])
    ws2 = wb.create_sheet("空白模板")
    _style_header(ws2, POS_HEADERS, [14, 14, 16, 12, 20, 12, 16])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_nos(tag_nos: dict[str, list[str]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "当前数据"
    _style_header(ws, NO_HEADERS, [12, 44])
    for no, epcs in tag_nos.items():
        first = True
        for epc in epcs or [""]:
            ws.append([no if first else "", epc])
            first = False
    ws2 = wb.create_sheet("空白模板")
    _style_header(ws2, NO_HEADERS, [12, 44])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_trolleys(trolleys: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "当前数据"
    _style_header(ws, TZ_HEADERS, [18, 14, 16])
    for t in trolleys:
        ws.append([t.get("name"), t.get("tz"), t.get("reader")])
    ws2 = wb.create_sheet("空白模板")
    _style_header(ws2, TZ_HEADERS, [18, 14, 16])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_table(title: str, headers: list[str], rows: list[list], widths: list[int] | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = (title or "导出")[:31]
    _style_header(ws, headers, widths or [18] * len(headers))
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
